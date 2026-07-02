"""Build the output workbook matching forwarder_detailed_dashboard.xlsx.

Sheet 1 "Detailed": one row per invoice (11 columns).
Sheet 2 "Summary": charges aggregated by Project No + Vendor.
"""
from __future__ import annotations

import io
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .schema import InvoiceRow

DETAILED_HEADERS = [
    "Project No", "Vendor", "Invoice No", "POD", "Type",
    "Freight (RM)", "Local Charges (RM)", "Port Storage (RM)",
    "Transport Charges (RM)", "Reimbursement (RM)", "Total (RM)",
]
SUMMARY_HEADERS = [
    "Project No", "Vendor", "Freight", "Local Charges",
    "Port Storage", "Transport Charges", "Reimbursement",
]

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_TOTAL_FONT = Font(bold=True)
_TOTAL_FILL = PatternFill("solid", fgColor="DDEBF7")
_MONEY_FMT = "#,##0.00"
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _style_header(ws, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER
        cell.border = _BORDER
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"


def _autosize(ws, ncols: int, minw: int = 10, maxw: int = 34) -> None:
    for c in range(1, ncols + 1):
        letter = get_column_letter(c)
        longest = max(
            (len(str(ws.cell(row=r, column=c).value or "")) for r in range(1, ws.max_row + 1)),
            default=minw,
        )
        ws.column_dimensions[letter].width = max(minw, min(maxw, longest + 2))


def _build_detailed(wb: Workbook, rows: list[InvoiceRow]) -> None:
    ws = wb.active
    ws.title = "Detailed"
    ws.append(DETAILED_HEADERS)

    money_cols = set(range(6, 12))  # columns F..K
    for row in rows:
        ws.append([
            row.project_no, row.vendor, row.invoice_no, row.pod, row.type,
            row.freight, row.local_charges, row.port_storage,
            row.transport_charges, row.reimbursement, row.total,
        ])

    # Totals row.
    total_row = ws.max_row + 1
    if rows:
        ws.cell(row=total_row, column=1, value="TOTAL")
        for c in money_cols:
            letter = get_column_letter(c)
            ws.cell(
                row=total_row, column=c,
                value=f"=SUM({letter}2:{letter}{total_row - 1})",
            )

    # Formatting.
    for r in range(2, ws.max_row + 1):
        is_total = r == total_row and bool(rows)
        for c in range(1, len(DETAILED_HEADERS) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = _BORDER
            if c in money_cols:
                cell.number_format = _MONEY_FMT
                cell.alignment = _CENTER
            else:
                cell.alignment = _LEFT
            if is_total:
                cell.font = _TOTAL_FONT
                cell.fill = _TOTAL_FILL

    _style_header(ws, len(DETAILED_HEADERS))
    _autosize(ws, len(DETAILED_HEADERS))


def _build_summary(wb: Workbook, rows: list[InvoiceRow]) -> None:
    ws = wb.create_sheet("Summary")
    ws.append(SUMMARY_HEADERS)

    # Aggregate by (project_no, vendor).
    agg: dict[tuple[str, str], list[float]] = defaultdict(lambda: [0.0, 0.0, 0.0, 0.0, 0.0])
    order: list[tuple[str, str]] = []
    for row in rows:
        key = (row.project_no or "(blank)", row.vendor)
        if key not in agg:
            order.append(key)
        bucket = agg[key]
        bucket[0] += row.freight
        bucket[1] += row.local_charges
        bucket[2] += row.port_storage
        bucket[3] += row.transport_charges
        bucket[4] += row.reimbursement

    for (project, vendor) in order:
        b = agg[(project, vendor)]
        ws.append([project, vendor, round(b[0], 2), round(b[1], 2),
                   round(b[2], 2), round(b[3], 2), round(b[4], 2)])

    money_cols = set(range(3, 8))  # C..G
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(SUMMARY_HEADERS) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = _BORDER
            if c in money_cols:
                cell.number_format = _MONEY_FMT
                cell.alignment = _CENTER
            else:
                cell.alignment = _LEFT

    _style_header(ws, len(SUMMARY_HEADERS))
    _autosize(ws, len(SUMMARY_HEADERS))


def build_workbook(rows: list[InvoiceRow]) -> bytes:
    """Return the .xlsx file as bytes."""
    wb = Workbook()
    _build_detailed(wb, rows)
    _build_summary(wb, rows)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
